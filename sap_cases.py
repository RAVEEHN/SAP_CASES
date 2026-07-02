#!/usr/bin/env python3
"""
SAP Cases Downloader
Fetches SAP support cases from me.sap.com, optionally filters, and emails a summary via Outlook.
Session is stored in ~/.sap_session/me_sap_cookies.json.
Runs on macOS and Windows.
"""

import json
import os
import platform
import requests
from datetime import datetime
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────

BASE         = "https://me.sap.com"
ODATA_PATH   = "/backend/raw/esrc/ReportingDataVerticle/odata/report/ODCases"
LANDING_URL  = f"{BASE}/app/cases"

SESSION_DIR  = Path.home() / ".sap_session"
COOKIES_FILE = SESSION_DIR / "me_sap_cookies.json"
STORAGE_FILE = SESSION_DIR / "me_sap_storage.json"

import sys as _sys
HERE = Path(_sys.executable).parent.resolve() if getattr(_sys, "frozen", False) else Path(__file__).parent.resolve()
CONFIG_FILE  = HERE / "config.json"

IS_WINDOWS = platform.system() == "Windows"

MONTHS_AGO = 6
PAGE_SIZE  = 100

EMAIL_FIELDS = ["CASE_NUMBER", "DESCRIPTION", "COMPONENT", "EXTERNAL_STATUS", "PRIO_TXT", "CREATION_DATE", "CHANGE_DATE"]

def _load_config() -> dict:
    if not CONFIG_FILE.exists():
        raise SystemExit(
            f"Config file not found: {CONFIG_FILE}\n"
            f'Create it with: {{"customer_number": "450162", "email_to": "you@sap.com"}}'
        )
    return json.loads(CONFIG_FILE.read_text())

_cfg           = _load_config()
EMAIL_TO       = _cfg["email_to"]
DAILY_CUSTOMER = _cfg["customer_number"]
SYSTEM_IDS     = set(_cfg.get("system_ids", []))  # empty = no filter
PRIORITIES     = set(_cfg.get("priorities", ["Very High", "High", "Medium", "Low"]))

GROUPBY_FIELDS = [
    "AGE", "CASE_ID", "CASE_NUMBER", "CASE_URL", "CHANGE_DATE",
    "CLOSING_DATE", "CLOSING_MONTH", "CLOSING_QUARTER", "CLOSING_WEEK", "CLOSING_YEAR",
    "COMPONENT", "COMPONENT_TXT", "CONF_TYPE", "CONTRACT_TYPE_SYS",
    "CREATION_DATE", "CREATION_MONTH", "CREATION_QUARTER", "CREATION_WEEK", "CREATION_YEAR",
    "CUSTOMER_ERP_ID", "DEPLOYMENT_TYPE", "DESCRIPTION", "ERROR_CAT_TXT", "ERROR_SUBCAT_TXT",
    "EXTERNAL_STATUS", "INIT_COMPONENT", "INIT_COMPONENT_TXT", "INIT_PRIO_TXT",
    "INSTALLATION_NAME", "INSTALLATION_NO", "IS_OPEN", "IS_REOPENED", "Id",
    "PARTNER_FLAG", "PARTNER_TXT", "PBUP_AC", "PBUP_ACTXTLG", "PRIO_LEVEL", "PRIO_TXT",
    "REPORTER_TYPE", "REQUIRES_ACTION_BY", "SOLUTION_AREA", "SOLUTION_AREA_TXT",
    "SOLUTION_SUB_AREA", "SOLUTION_SUB_AREA_TXT", "SOURCE_TXT",
    "SYSTEM_ID", "SYSTEM_NUMBER", "SYSTEM_PRODUCT", "SYSTEM_PRODUCT_TXT",
    "SYSTEM_TYPE", "SYSTEM_TYPE_TXT", "TOP_PRIO_TXT",
]

AGGREGATE_FIELDS = [
    "AVG_DAYS_CUSTOMER", "AVG_DAYS_SAP", "AVG_DAYS_TOTAL",
    "NUM_ACON", "NUM_CLOSED", "NUM_CLOSED_H", "NUM_CLOSED_L", "NUM_CLOSED_LAST6M",
    "NUM_CLOSED_M", "NUM_CLOSED_VH", "NUM_CONF", "NUM_HIGH", "NUM_LOW", "NUM_MED",
    "NUM_NON_PROD_TENANTS", "NUM_OPEN", "NUM_PROD_TENANTS", "NUM_REOPENINGS",
    "NUM_TOTAL", "NUM_TOTAL_LAST6M", "NUM_V_HIGH",
]


# ── Auth ──────────────────────────────────────────────────────────────────────

def _save_state(context):
    SESSION_DIR.mkdir(exist_ok=True)
    COOKIES_FILE.write_text(json.dumps(context.cookies(), indent=2))
    STORAGE_FILE.write_text(json.dumps(context.storage_state(), indent=2))


def _dismiss_cert_picker():
    try:
        import pyautogui, time
        time.sleep(3)
        pyautogui.press("enter")
        print("  Certificate picker 1 dismissed.")
        time.sleep(2)
        pyautogui.press("enter")
        print("  Certificate picker 2 dismissed.")
    except ImportError:
        print("  pyautogui not installed — skipping cert picker automation.")
    except Exception as e:
        print(f"  Could not dismiss cert picker: {e}")


def _find_system_browser() -> str | None:
    """Return path to installed Edge or Chrome on Windows, None otherwise."""
    candidates = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    ]
    for p in candidates:
        if Path(p).exists():
            return p
    return None


def _login_headful(playwright):
    import threading, time
    print("\nOpening browser for me.sap.com login...")
    print("Complete SSO login — browser closes automatically once done.\n")

    system_browser = _find_system_browser() if IS_WINDOWS else None
    if system_browser:
        browser = playwright.chromium.launch(headless=False, executable_path=system_browser)
    else:
        browser = playwright.chromium.launch(headless=False)

    context = browser.new_context()
    page    = context.new_page()
    threading.Thread(target=_dismiss_cert_picker, daemon=True).start()
    page.goto(LANDING_URL)
    print("Waiting for login (up to 5 minutes)...")
    page.wait_for_selector(
        "div[class*='case'], table, div[id*='app']",
        timeout=300_000,
    )
    time.sleep(3)
    _save_state(context)
    browser.close()
    print("Login successful — session saved.")


def _session_valid() -> bool:
    if not COOKIES_FILE.exists():
        return False
    try:
        s = get_session()
        resp = s.get(
            f"{BASE}{ODATA_PATH}",
            params={"$top": "1", "$apply": "groupby((CASE_NUMBER))"},
            timeout=15,
        )
        return resp.status_code == 200
    except Exception:
        return False


def ensure_session(force_relogin: bool = False):
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        raise SystemExit("playwright not installed. Run: pip install playwright && playwright install chromium")

    if not force_relogin and _session_valid():
        print("  Reusing saved me.sap.com session.")
        return

    with sync_playwright() as pw:
        _login_headful(pw)


# ── HTTP session ──────────────────────────────────────────────────────────────

def get_session() -> requests.Session:
    if not COOKIES_FILE.exists():
        raise SystemExit("No saved session. Run: python sap_cases.py --relogin")
    s = requests.Session()
    s.headers.update({
        "accept":           "application/json",
        "accept-language":  "en-US",
        "DNT":              "1",
        "User-Agent":       "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36 Edg/149.0.0.0",
        "x-requested-with": "XMLHttpRequest",
    })
    for c in json.loads(COOKIES_FILE.read_text()):
        s.cookies.set(c["name"], c["value"], domain=c.get("domain", "").lstrip("."))
    return s


def fetch_csrf(s: requests.Session) -> str:
    resp = s.get(
        f"{ODATA_PATH}?$top=1",
        headers={"x-csrf-token": "Fetch"},
    )
    token = resp.headers.get("x-csrf-token") or resp.headers.get("X-CSRF-Token")
    return token or ""


# ── Fetch cases ───────────────────────────────────────────────────────────────

def _encode_customer(customer_number: str) -> str:
    import base64, gzip
    return base64.b64encode(gzip.compress(customer_number.encode(), mtime=0)).decode().rstrip("=")


def _apply_param(customer_number: str) -> str:
    encoded   = _encode_customer(customer_number)
    groupby   = ",".join(GROUPBY_FIELDS)
    aggregate = ",".join(AGGREGATE_FIELDS)
    return (
        f"filter(CUSTOMER_ERP_ID eq '{encoded}' "
        f"and CREATION_MONTHS_AGO le {MONTHS_AGO})"
        f"/groupby(({groupby}),aggregate({aggregate}))"
    )


def fetch_all_cases(s: requests.Session, customer_number: str) -> list[dict]:
    url    = f"{BASE}{ODATA_PATH}"
    params = {
        "$apply": _apply_param(customer_number),
        "$top":   PAGE_SIZE,
        "$skip":  0,
    }

    all_cases = []
    total     = None

    while True:
        print(f"  Fetching records {params['$skip']+1}-{params['$skip']+PAGE_SIZE}"
              + (f" of {total}" if total else "") + " ...", flush=True)
        resp = s.get(url, params=params)

        if resp.status_code in (401, 403):
            return None

        resp.raise_for_status()
        data = resp.json()

        if total is None:
            total = data.get("@odata.count") or data.get("odata.count")

        batch = data.get("value", [])
        all_cases.extend(batch)

        if len(batch) < PAGE_SIZE:
            break
        params["$skip"] += PAGE_SIZE

    return all_cases


# ── Excel output ──────────────────────────────────────────────────────────────

def save_excel(cases: list[dict], customer_number: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise SystemExit("openpyxl not installed. Run: pip install openpyxl")

    output_file = HERE / f"sap_cases_{customer_number}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SAP Cases"

    headers = GROUPBY_FIELDS + AGGREGATE_FIELDS
    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.font       = Font(bold=True, color="FFFFFF")
        cell.fill       = PatternFill("solid", fgColor="1F4E79")
        cell.alignment  = Alignment(horizontal="center")

    for row_num, case in enumerate(cases, 2):
        for col, field in enumerate(headers, 1):
            val  = case.get(field, "")
            cell = ws.cell(row=row_num, column=col, value=val)
            if field == "CASE_NUMBER" and case.get("CASE_URL"):
                cell.hyperlink = case["CASE_URL"]
                cell.font      = Font(color="0563C1", underline="single")

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    ws.cell(row=len(cases) + 3, column=1,
            value=f"Downloaded: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    wb.save(output_file)
    print(f"  Saved {len(cases)} cases -> {output_file}")


# ── Filter ────────────────────────────────────────────────────────────────────

EXCLUDED_STATUSES = {"Confirmed", "Auto Confirmed"}

def filter_cases(cases: list[dict], open_only: bool = True) -> list[dict]:
    result = []
    for c in cases:
        if open_only and c.get("IS_OPEN") != "X":
            continue
        if c.get("EXTERNAL_STATUS") in EXCLUDED_STATUSES:
            continue
        if SYSTEM_IDS and c.get("SYSTEM_ID") not in SYSTEM_IDS:
            continue
        if c.get("PRIO_TXT") not in PRIORITIES:
            continue
        result.append(c)
    return result


# ── Email ─────────────────────────────────────────────────────────────────────

def _build_html_table(cases: list[dict], customer_number: str) -> str:
    header_cells = "".join(
        f'<th style="background:#1F4E79;color:#fff;padding:8px 12px;text-align:left">{f}</th>'
        for f in EMAIL_FIELDS
    )
    rows = []
    for c in cases:
        cells = []
        for f in EMAIL_FIELDS:
            val = c.get(f, "")
            if f == "CASE_NUMBER" and c.get("CASE_URL"):
                val = f'<a href="{c["CASE_URL"]}">{val}</a>'
            cells.append(f'<td style="padding:6px 12px;border-bottom:1px solid #ddd">{val}</td>')
        rows.append(f'<tr>{"".join(cells)}</tr>')

    table = (
        f'<table style="border-collapse:collapse;font-family:Arial,sans-serif;font-size:13px">'
        f'<thead><tr>{header_cells}</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody>'
        f'</table>'
    )
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    return (
        f'<p style="font-family:Arial,sans-serif;font-size:13px">'
        f'Open SAP cases for customer <b>{customer_number}</b> (excluding Confirmed / Auto Confirmed) - as of {ts} ({len(cases)} cases)</p>'
        f'{table}'
    )


def _send_email_windows(subject: str, html_body: str):
    try:
        import win32com.client
    except ImportError:
        raise SystemExit("pywin32 not installed. Run: pip install pywin32")
    outlook  = win32com.client.Dispatch("Outlook.Application")
    msg      = outlook.CreateItem(0)
    msg.To   = EMAIL_TO
    msg.Subject = subject
    msg.HTMLBody = html_body
    msg.Send()


def _send_email_macos(subject: str, html_body: str):
    import subprocess, tempfile, os
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".html", delete=False, encoding="utf-8")
    tmp.write(html_body)
    tmp.close()

    safe_subject  = subject.replace("\\", "\\\\").replace('"', '\\"')
    safe_tmp_path = tmp.name.replace("\\", "\\\\").replace('"', '\\"')

    script = f"""
set htmlBody to do shell script "cat " & quoted form of "{safe_tmp_path}"
tell application "Microsoft Outlook"
    set newMsg to make new outgoing message with properties {{subject:"{safe_subject}", content:htmlBody}}
    make new to recipient of newMsg with properties {{email address:{{name:"", address:"{EMAIL_TO}"}}}}
    send newMsg
end tell
"""
    try:
        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
        if result.returncode != 0:
            raise RuntimeError(f"Outlook send failed:\n{result.stderr}")
    finally:
        os.unlink(tmp.name)


def send_email_outlook(cases: list[dict], customer_number: str):
    subject   = f"Open Cases - {datetime.now().strftime('%Y-%m-%d')}"
    html_body = _build_html_table(cases, customer_number)

    if IS_WINDOWS:
        _send_email_windows(subject, html_body)
    else:
        _send_email_macos(subject, html_body)

    print(f"  Email sent to {EMAIL_TO} ({len(cases)} cases)")


# ── Scheduled task ────────────────────────────────────────────────────────────

def _manage_cron(install: bool):
    if IS_WINDOWS:
        _manage_task_scheduler(install)
    else:
        _manage_launchd(install)


def _manage_task_scheduler(install: bool):
    import subprocess, sys
    task_hourly = "SAPCasesErgon"
    task_logon  = "SAPCasesErgonLogon"
    script      = str(HERE / "sap_cases.py")
    binary      = str(Path(sys.executable))
    log         = HERE / "sap_cases.log"
    tr          = f'"{binary}" --scheduled >> "{log}" 2>&1'

    if not install:
        for task in (task_hourly, task_logon):
            subprocess.run(["schtasks", "/Delete", "/TN", task, "/F"], capture_output=True)
        print("Tasks removed.")
        return

    # Hourly 08:00-18:00
    cmd_hourly = [
        "schtasks", "/Create", "/F",
        "/TN", task_hourly,
        "/TR", tr,
        "/SC", "HOURLY",
        "/MO", "1",
        "/ST", "08:00",
        "/ET", "18:00",
        "/K",
        "/SD", datetime.now().strftime("%m/%d/%Y"),
    ]
    result = subprocess.run(cmd_hourly, capture_output=True, text=True)
    if result.returncode != 0:
        raise SystemExit(f"schtasks failed:\n{result.stderr}")

    # At logon — run as current user only (no admin needed)
    cmd_logon = [
        "schtasks", "/Create", "/F",
        "/TN", task_logon,
        "/TR", tr,
        "/SC", "ONLOGON",
        "/RU", os.environ.get("USERNAME", ""),
    ]
    result = subprocess.run(cmd_logon, capture_output=True, text=True)
    if result.returncode != 0:
        raise SystemExit(f"schtasks logon task failed:\n{result.stderr}")

    print(f"Registered Task Scheduler jobs — runs at logon and hourly 08:00-18:00.")
    print(f"Logs: {log}")


def _manage_launchd(install: bool):
    import subprocess, sys, stat
    plist_dir = Path.home() / "Library/LaunchAgents"
    log       = HERE / "sap_cases.log"
    binary    = str(Path(sys.executable))
    script    = str(HERE / "sap_cases.py")

    SLEEPWATCHER  = "/opt/homebrew/sbin/sleepwatcher"
    WAKEUP_SCRIPT = HERE / ".wakeup"
    label         = "com.user.sap-cases-ergon-sleepwatcher"

    if not install:
        plist = plist_dir / f"{label}.plist"
        subprocess.run(["launchctl", "unload", str(plist)], capture_output=True)
        if plist.exists():
            plist.unlink()
        for old in ("com.user.sap-cases-ergon-watcher", "com.user.sap-cases-ergon-login",
                    "com.user.sap-cases-ergon-unlock", "com.user.sap-cases-ergon"):
            p = plist_dir / f"{old}.plist"
            subprocess.run(["launchctl", "unload", str(p)], capture_output=True)
            if p.exists():
                p.unlink()
        print("Jobs removed.")
        return

    if not Path(SLEEPWATCHER).exists():
        raise SystemExit(f"sleepwatcher not found at {SLEEPWATCHER}. Run: brew install sleepwatcher")

    wakeup_content = (
        f"#!/bin/sh\n"
        f"pgrep -u \"$USER\" -x Finder > /dev/null 2>&1 || exit 0\n"
        f"{binary} {script} --scheduled >> {log} 2>&1 &\n"
    )
    WAKEUP_SCRIPT.write_text(wakeup_content)
    WAKEUP_SCRIPT.chmod(WAKEUP_SCRIPT.stat().st_mode | stat.S_IXUSR)

    plist_dir.mkdir(parents=True, exist_ok=True)
    plist = plist_dir / f"{label}.plist"
    plist.write_text(f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN"
  "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
<plist version="1.0">
<dict>
    <key>Label</key><string>{label}</string>
    <key>ProgramArguments</key>
    <array>
        <string>{SLEEPWATCHER}</string>
        <string>-V</string>
        <string>-w</string>
        <string>{WAKEUP_SCRIPT}</string>
    </array>
    <key>RunAtLoad</key><true/>
    <key>KeepAlive</key><true/>
    <key>StandardOutPath</key><string>{log}</string>
    <key>StandardErrorPath</key><string>{log}</string>
</dict>
</plist>""")
    subprocess.run(["launchctl", "unload", str(plist)], capture_output=True)
    result = subprocess.run(["launchctl", "load", str(plist)], capture_output=True, text=True)
    if result.returncode != 0:
        raise SystemExit(f"launchctl load failed:\n{result.stderr}")

    subprocess.Popen([binary, script, "--customer", DAILY_CUSTOMER, "--email"],
                     stdout=open(log, "a"), stderr=subprocess.STDOUT)

    print("Registered — runs on login and each screen wake/unlock.")
    print(f"Wakeup script: {WAKEUP_SCRIPT}")
    print(f"Logs: {log}")


LAST_RUN_FILE = HERE / ".last_run"


def _already_ran_today() -> bool:
    if LAST_RUN_FILE.exists():
        return LAST_RUN_FILE.read_text().strip() == datetime.now().strftime("%Y-%m-%d")
    return False


def _mark_ran_today():
    LAST_RUN_FILE.write_text(datetime.now().strftime("%Y-%m-%d"))


# ── Main ──────────────────────────────────────────────────────────────────────

def main(force_relogin: bool = False, customer_number: str = "", send_email: bool = False, scheduled: bool = False):
    customer_number = customer_number.zfill(10)
    print(f"\n-- {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} -- customer: {customer_number} --")

    if scheduled and send_email and _already_ran_today():
        print("  Already ran today -- skipping.")
        return

    ensure_session(force_relogin=force_relogin)
    s     = get_session()
    cases = fetch_all_cases(s, customer_number)

    if cases is None:
        print("  Session expired -- re-authenticating...")
        ensure_session(force_relogin=True)
        s     = get_session()
        cases = fetch_all_cases(s, customer_number)

    if cases is None:
        raise SystemExit("Failed to fetch cases even after re-authentication.")

    print(f"  Total cases fetched: {len(cases)}")
    save_excel(cases, customer_number)

    if send_email:
        filtered = filter_cases(cases, open_only=True)
        print(f"  Filtered to {len(filtered)} open cases (excluding Confirmed / Auto Confirmed)")
        if filtered:
            send_email_outlook(filtered, customer_number)
        else:
            print("  No open cases -- email skipped.")
        _mark_ran_today()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Download SAP support cases to Excel")
    parser.add_argument("--relogin",   action="store_true", help="Force fresh browser login")
    parser.add_argument("--customer",  default="", help="Customer number (overrides config.json)")
    parser.add_argument("--email",     action="store_true", help="Filter and email open cases via Outlook")
    parser.add_argument("--scheduled", action="store_true", help="Run from scheduler — skips if already ran today")

    sub = parser.add_subparsers(dest="cmd")
    sub.add_parser("setup-cron",  help=f"Register scheduled job for customer {DAILY_CUSTOMER}")
    sub.add_parser("remove-cron", help="Unregister scheduled job")

    args = parser.parse_args()

    if args.cmd == "setup-cron":
        _manage_cron(install=True)
    elif args.cmd == "remove-cron":
        _manage_cron(install=False)
    else:
        customer = args.customer or DAILY_CUSTOMER
        main(force_relogin=args.relogin, customer_number=customer, send_email=True, scheduled=args.scheduled)
